import openpyxl
import time
import re
import os
try:
    import undetected_chromedriver as uc
    USE_UC = True
except ImportError:
    USE_UC = False
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from difflib import SequenceMatcher


def normalize(text):
    text = str(text).lower()
    # Убираем теги года и версии: [2025], (NEW for 2025!), 2026, v2, etc.
    text = re.sub(r"\[\d{4}\]", "", text)
    text = re.sub(r"\(new[^)]*\)", "", text)
    text = re.sub(r"20\d{2}", "", text)
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()


def init_browser():
    print("Запускаем браузер...")
    if USE_UC:
        print("Используем undetected-chromedriver (обход Cloudflare)")
        options = uc.ChromeOptions()
        options.add_argument("--window-size=1400,900")
        options.page_load_strategy = "eager"
        driver = uc.Chrome(options=options, version_main=145)
    else:
        print("undetected-chromedriver не установлен, используем обычный Chrome")
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        from webdriver_manager.chrome import ChromeDriverManager
        options = Options()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        options.add_argument("--window-size=1400,900")
        options.add_argument(
            "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
        )
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def search_udemy(driver, topic):
    topic_norm = normalize(topic)
    words = topic.split()

    # Пробуем полный запрос, потом укороченный
    queries = [topic]
    if len(words) > 6:
        queries.append(" ".join(words[:6]))
    if len(words) > 4:
        queries.append(" ".join(words[:4]))

    for q_idx, query in enumerate(queries):
        print(f"  Поиск: «{query[:70]}»")
        try:
            encoded = re.sub(r'\s+', '+', query.strip()).replace('&', '%26').replace('#', '%23')
            url = f"https://www.udemy.com/courses/search/?q={encoded}&sort=relevance"
            driver.get(url)

            # Ждём появления карточек
            try:
                WebDriverWait(driver, 7).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/course/']"))
                )
            except Exception:
                print("  Страница не загрузилась, пробуем дальше...")
                time.sleep(1)
                continue

            time.sleep(0.2)  # минимальная пауза

            # Берём только первые 3 ссылки /course/
            links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/course/']")
            seen = set()
            results = []
            for el in links:
                href = el.get_attribute("href") or ""
                href = re.sub(r'\?.*$', '', href).rstrip('/')
                if '/course/' not in href or href in seen:
                    continue
                seen.add(href)
                title = el.get_attribute("title") or el.text.strip()
                if not title:
                    slug = href.split('/course/')[-1].rstrip('/')
                    title = slug.replace('-', ' ')
                results.append((title, href))
                if len(results) >= 5:
                    break

            if not results:
                print("  → результатов нет")
                continue

            # Порог: для полного запроса мягче, для укороченных — строже
            # Укороченный запрос может поймать мусор → нужен высокий порог
            if q_idx == 0:
                # Полный запрос: порог по длине названия
                if len(words) <= 3:
                    threshold = 0.90
                elif len(words) <= 5:
                    threshold = 0.84
                else:
                    threshold = 0.80
            else:
                # Укороченный запрос: всегда строго
                threshold = 0.90
            for title, href in results:
                score = similarity(normalize(title), topic_norm)
                print(f"   [{score:.2f}] {title[:65]}")
                if score >= threshold:
                    print(f" → найден ({score:.2f}): {href}")
                    return href

            print(f" → нет точного совпадения в топ-3")
        except Exception as e:
            print(f"  Ошибка: {type(e).__name__}: {str(e)[:80]}")
            time.sleep(3)

    return None


def process_excel(file):
    if not os.path.isfile(file):
        print(f"Файл не найден: {file}")
        return

    print(f"\nОткрываем файл: {file}")
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    driver = init_browser()
    updated = 0
    processed = 0
    requests_count = 0  # счётчик запросов для перезапуска браузера

    # Открываем Udemy чтобы пройти Cloudflare вручную
    print("" + "="*50)
    print("Открываем Udemy...")
    print("Если появится Cloudflare проверка — нажми чекбокс вручную!")
    print("У тебя есть 12 секунд.")
    print("="*50)
    driver.get("https://www.udemy.com/")
    time.sleep(12)
    print("Продолжаем...")

    try:
        for r in range(2, ws.max_row + 1):
            topic = ws.cell(row=r, column=1).value
            link_cell = ws.cell(row=r, column=13)

            existing = str(link_cell.value).strip() if link_cell.value else ""
            if existing.startswith("http"):
                continue
            if not topic:
                continue

            topic = str(topic).strip()
            processed += 1
            print(f"\n{r} | {topic}")

            # Профилактический перезапуск каждые 30 курсов
            requests_count += 1
            if requests_count % 50 == 0:
                print(" Перезапускаем браузер (профилактика)...")
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(2)
                driver = init_browser()
                driver.get("https://www.udemy.com/")
                time.sleep(5)

            # Запуск поиска с перезапуском при падении браузера
            try:
                link = search_udemy(driver, topic)
            except Exception as e:
                err = str(e)
                if any(x in err for x in ["invalid session", "browser has closed", "disconnected", "InvalidSessionId"]):
                    print(" Браузер упал, перезапускаем...")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    time.sleep(3)
                    driver = init_browser()
                    driver.get("https://www.udemy.com/")
                    time.sleep(5)
                    try:
                        link = search_udemy(driver, topic)
                    except Exception:
                        link = None
                else:
                    print(f" Ошибка: {err[:80]}")
                    link = None

            if link:
                link_cell.value = link
                updated += 1
            else:
                link_cell.value = "Не найдено"
                print(" → курс не найден")

            # Сохраняем сразу если нашли ссылку, иначе каждые 5 строк
            if link or processed % 5 == 0:
                wb.save(file)
                print(f"Файл сохранён (найдено: {updated})")

    finally:
        driver.quit()
        print(f"\nГотово. Найдено: {updated}")


if __name__ == "__main__":
    FILE = "final_dataset_ready.xlsx"
    print("Папка:", os.getcwd())
    print("Файл:", os.path.abspath(FILE))
    print("Существует:", os.path.isfile(FILE))
    print("-" * 50)
    process_excel(FILE)