import openpyxl
import time
import random
import concurrent.futures

import undetected_chromedriver as uc

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


FILE = "final_dataset_ready.xlsx"

MAX_WORKERS = 4   # больше не ставь на Mac


# -------------------------------
# создание драйвера
# -------------------------------

def create_driver():

    options = uc.ChromeOptions()

    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    driver = uc.Chrome(options=options)

    driver.set_page_load_timeout(30)

    return driver


# -------------------------------
# парсинг страницы
# -------------------------------

def parse_course(row_data):

    row, url = row_data

    for attempt in range(3):

        driver = None

        try:

            time.sleep(random.uniform(1, 3))

            driver = create_driver()

            driver.get(url)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            title = driver.title

            description = ""

            try:
                description = driver.find_element(By.TAG_NAME, "meta").get_attribute("content")
            except:
                pass

            driver.quit()

            print(f"Получено: {row}")

            return row, title, description

        except Exception as e:

            print("Ошибка:", e)

            if driver:
                driver.quit()

            time.sleep(2)

    return row, None, None


# -------------------------------
# основной запуск
# -------------------------------

def run():

    wb = openpyxl.load_workbook(FILE)

    ws = wb.active

    tasks = []

    for row in range(2, ws.max_row + 1):

        url = ws.cell(row=row, column=13).value

        if url:
            tasks.append((row, url))


    results = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:

        futures = [executor.submit(parse_course, t) for t in tasks]

        for f in concurrent.futures.as_completed(futures):

            result = f.result()

            if result:
                results.append(result)


    for row, title, description in results:

        if title:
            ws.cell(row=row, column=1).value = title

        if description:
            ws.cell(row=row, column=10).value = description


    wb.save(FILE)

    print("Готово")


if __name__ == "__main__":

    run()