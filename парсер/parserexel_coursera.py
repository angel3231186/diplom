# parserexel.py
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import os

def init_browser():
    """Инициализирует один браузер на весь скрипт"""
    print("Запускаем браузер... (один раз)")
    options = Options()
    options.page_load_strategy = 'eager'
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36')

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    return driver


def get_best_course(driver, topic: str, wait: WebDriverWait) -> str | None:
    """Использует существующий браузер"""
    if not topic or not isinstance(topic, str):
        return None

    topic = topic.strip()
    max_attempts = 2

    for attempt in range(1, max_attempts + 1):
        try:
            print(f"  Попытка {attempt}/{max_attempts} → {topic[:70]}{'...' if len(topic)>70 else ''}")

            query = topic.replace(' ', '+').replace('&', '%26')
            driver.get(f'https://www.coursera.org/search?query={query}')

            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/learn/']")))

            elements = driver.find_elements(By.CSS_SELECTOR, "a[href*='/learn/']")
            if elements:
                href = elements[0].get_attribute('href')
                clean = href.split('?')[0].rstrip('/')
                print(f"     → {clean}")
                return clean

            print("     Элементы не найдены")

        except Exception as e:
            print(f"     Ошибка: {type(e).__name__} – {str(e)[:100]}...")

        if attempt < max_attempts:
            time.sleep(3)   

    print("     → Не удалось найти после всех попыток")
    return None


def process_excel(input_file: str):
    if not os.path.isfile(input_file):
        print(f"Файл не найден: {input_file}")
        return

    print(f"\nОткрываем: {input_file}")
    try:
        wb = load_workbook(input_file, data_only=True)
        sheet_name = "final_dataset_ready"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        print(f"Лист: {ws.title}")

        driver = init_browser()
        wait = WebDriverWait(driver, 25)  # можно уменьшить до 15–20

        updated = 0
        skipped = 0
        processed = 0

        for r in range(2, ws.max_row + 1):
            title = ws.cell(row=r, column=1).value
            if not title:
                continue

            link_cell = ws.cell(row=r, column=13)
            if link_cell.value and str(link_cell.value).strip():
                skipped += 1
                continue

            title_str = str(title).strip()
            print(f" {r:4d} | {title_str[:68]}{'...' if len(title_str)>68 else ''}")

            link = get_best_course(driver, title_str, wait)

            if link:
                link_cell.value = link
                updated += 1
                print(f"     Добавлена ссылка (обновлено: {updated})")
            else:
                link_cell.value = "Не найдено"
                print("     Не найдено")

            processed += 1

            # Сохранение после каждой строки
            try:
                wb.save(input_file)
                print(f"     Сохранено после строки {r}")
            except Exception as save_err:
                print(f"     Не удалось сохранить: {save_err}")

            time.sleep(0.8)  # ← уменьшено! (можно даже 0.8–1.0)

        driver.quit()  # закрываем браузер только в конце

        print("\n" + "═" * 70)
        print("Обработка завершена")
        print(f"  Добавлено ссылок   : {updated}")
        print(f"  Пропущено          : {skipped}")
        print(f"  Обработано пустых  : {processed}")
        print(f"  Всего строк        : {ws.max_row - 1}")
        print("═" * 70)
        print(f"Файл обновлён: {os.path.abspath(input_file)}")

    except Exception as e:
        print(f"\nКритическая ошибка: {type(e).__name__}: {e}")
        if 'driver' in locals():
            driver.quit()

if __name__ == "__main__":
    FILE = "final_dataset_ready.xlsx"
    print("Папка          :", os.getcwd())
    print("Файл           :", os.path.abspath(FILE))
    print("Существует?    :", os.path.isfile(FILE))
    print("-" * 60)
    process_excel(FILE)
