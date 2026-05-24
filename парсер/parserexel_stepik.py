import requests
import openpyxl
import time
import os
import re
from difflib import SequenceMatcher

KEYWORDS = [
    "python", "java", "javascript", "js", "kotlin", "swift", "c\\+\\+", "c#",
    "ruby", "go", "rust", "php", "scala", "matlab", "fortran",
    "django", "flask", "fastapi", "react", "angular", "vue", "node",
    "sql", "postgresql", "mysql", "mongodb", "redis", "clickhouse",
    "docker", "kubernetes", "linux", "git", "github",
    "tensorflow", "pytorch", "numpy", "pandas", "sklearn",
    "grafana", "prometheus", "kafka", "spark", "hadoop", "flink",
    "excel", "power bi", "tableau", "1с", "appmetrica",
    "selenium", "pytest", "unittest",
    "android", "ios", "flutter", "unity",
    "html", "css",
]

# Названия которые слишком короткие/общие — искать бессмысленно
TOO_GENERIC = {
    "анализ данных", "машинное обучение", "субд", "большие данные",
    "нейронные сети", "python", "pandas", "nosql", "основы sql",
    "анализ данных в r", "power bi", "базовая аналитика",
    "корпоративные финансы", "собеседование аналитика",
}


def is_english(text):
    """Проверяет что название преимущественно на английском"""
    clean = re.sub(r"[^a-zA-Zа-яА-Я]", "", text)
    if not clean:
        return False
    en = sum(1 for c in clean if c.isascii())
    return en / len(clean) > 0.7


def extract_keywords(text):
    text_lower = text.lower()
    found = set()
    for kw in KEYWORDS:
        if re.search(kw, text_lower):
            found.add(kw)
    return found


def keywords_match(topic_norm, title_norm):
    topic_kw = extract_keywords(topic_norm)
    if not topic_kw:
        return True
    title_kw = extract_keywords(title_norm)
    return topic_kw.issubset(title_kw)


def normalize(text):
    text = str(text).lower()
    text = re.sub(r"[-—]\s*stepik\s*$", "", text)
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_for_search(text):
    """Очищает название для поиска — убирает спецсимволы, заменяет на пробелы"""
    text = re.sub(r"[-—]\s*stepik\s*$", "", text, flags=re.IGNORECASE).strip()
    # Точки между словами → пробел
    text = re.sub(r"(?<=\w)\.(?=\s)", " ", text)
    # Плюс → пробел
    text = text.replace("+", " ")
    # Слеш → пробел
    text = re.sub(r"[/\\|]", " ", text)
    # Убираем лишние пробелы
    text = re.sub(r"\s+", " ", text).strip()
    return text


def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()


def shorten_query(topic):
    topic_clean = clean_for_search(topic)
    variants = []

    # Дефис → пробел
    dehyphen = re.sub(r"\s+", " ", topic_clean.replace("-", " ")).strip()
    if dehyphen != topic_clean:
        variants.append(dehyphen)

    # Без скобок
    no_brackets = re.sub(r"[\(\[\{][^\)\]\}]*[\)\]\}]", " ", topic_clean)
    no_brackets = re.sub(r"\s+", " ", no_brackets).strip()
    if no_brackets != topic_clean:
        variants.append(no_brackets)
        variants.append(no_brackets.replace("-", " "))

    # До двоеточия
    if ":" in topic_clean:
        before_colon = topic_clean.split(":")[0].strip()
        after_colon = topic_clean.split(":", 1)[1].strip()
        if before_colon:
            variants.append(before_colon)
            variants.append(before_colon.replace("-", " "))
        if after_colon and len(after_colon) > 5:
            variants.append(after_colon)

    # Без первого слова
    words = topic_clean.split()
    if len(words) > 3:
        variants.append(" ".join(words[1:]))
        variants.append(" ".join(words[1:]).replace("-", " "))

    # Первые N слов
    for n in [5, 4, 3]:
        if len(words) >= n:
            chunk = " ".join(words[:n])
            if chunk not in variants:
                variants.append(chunk)
            chunk_dh = chunk.replace("-", " ")
            if chunk_dh not in variants:
                variants.append(chunk_dh)

    # Уникальные редкие слова (длиннее 5 букв, не стоп-слова)
    STOPWORDS = {"курс", "нуля", "основы", "часть", "модуль", "введение",
                 "практика", "полный", "быстрый", "python", "данных",
                 "анализ", "обучение", "машинное", "stepik", "онлайн"}
    rare_words = [w for w in words
                  if len(w) > 5
                  and w.lower() not in STOPWORDS
                  and not w.isdigit()]
    if rare_words:
        rare_words_sorted = sorted(rare_words, key=len, reverse=True)
        query_rare = " ".join(rare_words_sorted[:3])
        if query_rare not in variants:
            variants.append(query_rare)

    # Паттерн Сенаторова: "X в Python. Математика машинного обучения"
    # → ищем как "Математика Data Science с нуля: X Python"
    import re as _re
    ml_match = _re.search(r"математик", topic_clean.lower())
    ds_keywords = ["разложение", "регрессия", "регуляризация", "градиентный",
                   "сингулярное", "переобучение", "алгебра", "вероятностей"]
    if ml_match or any(kw in topic_clean.lower() for kw in ds_keywords):
        core = topic_clean.split(".")[0].strip()  # берём первую часть до точки
        # Берём только первые 3 слова core для короткого запроса
        core_short = " ".join(core.split()[:4])
        senatorov_q = "Математика Data Science с нуля " + core_short
        variants.append(senatorov_q)

    return list(dict.fromkeys(v for v in variants if v and len(v) > 3))


def query_stepik(search_term, topic_norm, threshold=0.95):
    try:
        url = "https://stepik.org/api/courses"
        params = {"search": search_term, "page": 1}
        r = requests.get(url, params=params, timeout=10)
        data = r.json()
        courses = data.get("courses", [])

        if not courses:
            return None, 0.0

        best_link = None
        best_score = 0.0

        for course in courses:
            title_norm = normalize(course["title"])
            score = similarity(title_norm, topic_norm)

            if not keywords_match(topic_norm, title_norm):
                print(f"   [{score:.2f}] ✗ {title_norm}")
                continue

            print(f"   [{score:.2f}] {title_norm}")

            if score > best_score:
                best_score = score
                best_link = f"https://stepik.org/course/{course['id']}"

        if best_score == 1.0:
            print(f" → точное совпадение: {best_link}")
            return best_link, 1.0

        if best_score >= threshold:
            print(f" → лучшее совпадение ({best_score:.2f}): {best_link}")
            return best_link, best_score

        return None, best_score

    except Exception as e:
        print("Ошибка запроса:", e)
        return None, 0.0


def search_stepik_best(topic):
    if not topic:
        return None

    topic_norm = normalize(topic)

    # Пропускаем слишком общие названия
    if topic_norm in TOO_GENERIC:
        print(" → слишком общее название, пропускаем")
        return None

    # Пропускаем преимущественно английские
    if is_english(topic):
        print(" → английское название, Stepik не поддерживает — пропускаем")
        return None

    # 1. Основной поиск (очищенное название)
    clean = clean_for_search(topic)
    print(f" Поиск: «{clean}»")
    link, score = query_stepik(clean, topic_norm)
    if link:
        return link

    # 2. Резервные варианты
    for short in shorten_query(topic):
        print(f" Резервный поиск: «{short}»")
        time.sleep(0.2)
        link, score = query_stepik(short, topic_norm, threshold=0.95)
        if link:
            return link

    return None


def process_excel(file):
    print("\nОткрываем файл:", file)
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    for r in range(2, ws.max_row + 1):
        topic = ws.cell(row=r, column=1).value
        link_cell = ws.cell(row=r, column=13)

        if link_cell.value:
            continue
        if not topic:
            continue

        topic = str(topic).strip()
        print(f"\n{r} | {topic}")

        link = search_stepik_best(topic)

        if link:
            link_cell.value = link
        else:
            link_cell.value = "Не найдено"
            print(" → курс не найден")

        wb.save(file)
        print("Файл сохранён")
        time.sleep(0.3)

    print("\nГотово")


if __name__ == "__main__":
    FILE = "final_dataset_ready.xlsx"
    print("Папка:", os.getcwd())
    print("Файл:", os.path.abspath(FILE))
    print("Существует:", os.path.isfile(FILE))
    print("-" * 50)

    if os.path.isfile(FILE):
        process_excel(FILE)
    else:
        print("Файл не найден!")